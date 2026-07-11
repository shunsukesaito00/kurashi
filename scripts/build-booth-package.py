#!/usr/bin/env python3
"""BOOTH運営者向け素材 ZIP（xlsx + PDFマニュアル + サムネイル）を生成する。

要: pip install openpyxl fpdf2
サムネイル PNG 生成には node + playwright（scripts/）が必要。
"""
from __future__ import annotations

import subprocess
import sys
import zipfile
from pathlib import Path

from fpdf import FPDF
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
PRODUCTS = ROOT / "products"
XLSX = PRODUCTS / "tedori-kakei-template.xlsx"
PDF = PRODUCTS / "manual.pdf"
THUMBNAIL = PRODUCTS / "booth-thumbnail.png"
ZIP_OUT = PRODUCTS / "tedori-kakei-booth.zip"
GENERATE_SCRIPT = ROOT / "scripts" / "generate-spreadsheet-template.py"
THUMBNAIL_SCRIPT = ROOT / "scripts" / "generate-booth-thumbnail.mjs"

FONT_CANDIDATES = [
    "/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc",
    "/System/Library/Fonts/Hiragino Sans GB.ttc",
    "/Library/Fonts/Arial Unicode.ttf",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJKjp-Regular.otf",
    "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
]


def find_japanese_font() -> Path:
    for candidate in FONT_CANDIDATES:
        path = Path(candidate)
        if path.exists():
            return path
    raise SystemExit(
        "日本語フォントが見つかりません。"
        " macOS では標準フォント、Linux では fonts-noto-cjk 等をインストールしてください。"
    )


def generate_xlsx() -> None:
    subprocess.run([sys.executable, str(GENERATE_SCRIPT)], check=True)


def ensure_thumbnail() -> Path:
    if not THUMBNAIL.exists():
        subprocess.run(["node", str(THUMBNAIL_SCRIPT)], check=True, cwd=THUMBNAIL_SCRIPT.parent)
    if not THUMBNAIL.exists():
        raise SystemExit(
            f"サムネイルが見つかりません: {THUMBNAIL}\n"
            " node scripts/generate-booth-thumbnail.mjs を実行してください。"
        )
    return THUMBNAIL


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_readme_content(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["README"]

    rows = []
    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=4, values_only=True):
        rows.append([_text(c) for c in row])

    title = rows[0][0]
    disclaimers = [rows[i][0].lstrip("・") for i in range(3, 8) if rows[i][0]]

    usage: list[tuple[str, str]] = []
    i = 9
    while i < len(rows) and rows[i][0] != "検証用テストケース":
        if rows[i][0] and rows[i][1]:
            usage.append((rows[i][0], rows[i][1]))
        i += 1

    test_cases: list[tuple[str, str, str, str]] = []
    if i < len(rows) and rows[i][0] == "検証用テストケース":
        i += 1
        if i < len(rows):
            headers = rows[i][:4]
            i += 1
            while i < len(rows) and rows[i][0] and rows[i][0] != "無料Webツール":
                case = rows[i][:4]
                if len(case) >= 3 and case[2]:
                    test_cases.append(tuple(case))
                i += 1

    web_url = ""
    web_note = ""
    while i < len(rows):
        if rows[i][0] == "無料Webツール":
            if i + 1 < len(rows):
                web_url = rows[i + 1][0]
            if i + 2 < len(rows):
                web_note = rows[i + 2][0]
            break
        i += 1

    wb.close()
    return {
        "title": title,
        "disclaimers": disclaimers,
        "usage": usage,
        "test_cases": test_cases,
        "web_url": web_url,
        "web_note": web_note,
    }


class ManualPDF(FPDF):
    def __init__(self, font_path: Path):
        super().__init__()
        self.font_path = font_path
        self.add_font("JP", "", str(font_path))
        self.set_auto_page_break(auto=True, margin=18)
        self.set_margins(18, 18, 18)

    def section(self, heading: str) -> None:
        self.ln(4)
        self.set_font("JP", size=12)
        self.multi_cell(0, 8, heading)
        self.ln(2)

    def body(self, text: str, size: int = 10) -> None:
        self.set_font("JP", size=size)
        self.multi_cell(0, 6, text)
        self.ln(1)

    def bullet(self, text: str) -> None:
        self.body(f"・{text}")


def format_yen(value: str) -> str:
    try:
        n = int(float(value))
        return f"{n:,}円"
    except ValueError:
        return value


def build_manual_pdf(content: dict, out_path: Path, font_path: Path) -> None:
    pdf = ManualPDF(font_path)
    pdf.add_page()

    pdf.set_font("JP", size=16)
    pdf.multi_cell(0, 10, content["title"])
    pdf.ln(2)
    pdf.body("くらしの計算室 BOOTH商品 — 使い方マニュアル（簡易版）", size=9)

    pdf.section("同梱ファイル")
    pdf.bullet("tedori-kakei-template.xlsx — 手取り試算・家計サマリ等の6シート入りテンプレート")
    pdf.bullet("manual.pdf — 本マニュアル")
    pdf.body("※ booth-thumbnail.png はBOOTH出品用サムネイル（運営者向け）です。購入者への配布物には含めません。", size=9)

    pdf.section("免責事項")
    for item in content["disclaimers"]:
        pdf.bullet(item)

    pdf.section("使い方")
    for title, detail in content["usage"]:
        pdf.body(title, size=10)
        pdf.body(detail, size=9)

    pdf.section("検証用テストケース")
    pdf.body("手取り試算が正しく動作しているか、次の代表例で確認できます。", size=9)
    for case in content["test_cases"]:
        salary, bonus, net, note = case
        pdf.body(
            f"月収 {format_yen(salary)} / ボーナス {format_yen(bonus)}"
            f" → 手取り {format_yen(net)}（{note}）",
            size=9,
        )

    if content["web_url"]:
        pdf.section("無料Webツール")
        pdf.body(content["web_url"], size=9)
        if content["web_note"]:
            pdf.body(content["web_note"], size=9)

    pdf.section("シート一覧")
    sheets = [
        "README — 免責・使い方・検証用テストケース",
        "手取り試算 — 月収・ボーナスから手取りを概算",
        "手取り比較 — 最大3パターンの比較",
        "家計サマリ — 固定費・変動費と自由に使える額",
        "積立メモ — 複利積立シミュレーション",
        "年間俯瞰 — 12か月の手取り・支出・貯蓄",
    ]
    for line in sheets:
        pdf.bullet(line)

    pdf.output(str(out_path))


def create_zip(paths: list[Path], zip_path: Path) -> None:
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in paths:
            zf.write(path, arcname=path.name)


def main() -> None:
    font_path = find_japanese_font()
    generate_xlsx()
    if not XLSX.exists():
        raise SystemExit(f"xlsx が見つかりません: {XLSX}")

    thumbnail = ensure_thumbnail()
    content = extract_readme_content(XLSX)
    build_manual_pdf(content, PDF, font_path)
    create_zip([XLSX, PDF, thumbnail], ZIP_OUT)

    print(f"OK: {ZIP_OUT.relative_to(ROOT)}")
    for path in (XLSX, PDF, thumbnail):
        print(f"  - {path.name}")


if __name__ == "__main__":
    main()
