#!/usr/bin/env python3
"""BOOTH向け手取り試算 .xlsx テンプレートを生成する。要: pip install openpyxl"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "products" / "tedori-kakei-template.xlsx"

SALARY_DED = (
    "=IF({c}5<=1625000,550000,"
    "IF({c}5<=1800000,{c}5*0.4-100000,"
    "IF({c}5<=3600000,{c}5*0.3+80000,"
    "IF({c}5<=6600000,{c}5*0.2+440000,"
    "IF({c}5<=8500000,{c}5*0.1+1100000,"
    "1950000)))))"
)

INCOME_TAX = (
    "=IF({c}11<=0,0,"
    "IF({c}11<=1950000,{c}11*0.05,"
    "IF({c}11<=3300000,{c}11*0.1-97500,"
    "IF({c}11<=6950000,{c}11*0.2-427500,"
    "IF({c}11<=9000000,{c}11*0.23-636000,"
    "IF({c}11<=18000000,{c}11*0.33-1536000,"
    "IF({c}11<=40000000,{c}11*0.4-2796000,"
    "{c}11*0.45-4796000)))))))*1.021"
)


def apply_net_formulas(ws, col: str, input_row_salary: int = 3, input_row_bonus: int = 4):
    c = col
    r = input_row_salary
    b = input_row_bonus
    ws[f"{c}5"] = f"={c}{r}*12+{c}{b}"
    ws[f"{c}6"] = f"=ROUND({c}{r}*0.0495,0)"
    ws[f"{c}7"] = f"=ROUND({c}{r}*0.0915,0)"
    ws[f"{c}8"] = f"=ROUND({c}{r}*0.0055,0)"
    ws[f"{c}9"] = f"=({c}6+{c}7+{c}8)*12+{c}{b}*0.1465"
    ws[f"{c}10"] = SALARY_DED.format(c=c)
    ws[f"{c}11"] = f"=MAX(0,{c}5-{c}10-{c}9-480000)"
    ws[f"{c}12"] = INCOME_TAX.format(c=c)
    ws[f"{c}13"] = f"=MAX(0,{c}5-{c}10-{c}9-430000)"
    ws[f"{c}14"] = f"={c}13*0.1+5000"
    ws[f"{c}15"] = f'=IF({c}5=0,0,ROUND({c}12/12*({c}{r}*12/{c}5),0))'
    ws[f"{c}16"] = f"=ROUND({c}14/12,0)"
    ws[f"{c}17"] = f"={c}{r}-{c}6-{c}7-{c}8-{c}15-{c}16"


def style_header(ws, title: str):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "※概算です。健保・扶養・住民税の開始時期等は反映しません。給与明細・勤務先でご確認ください。"
    ws["A2"].font = Font(size=9, color="666666")
    ws.column_dimensions["A"].width = 28
    for col in "BCD":
        ws.column_dimensions[col].width = 16


def style_yen(ws, cells):
    for ref in cells:
        ws[ref].number_format = "#,##0"


def build_readme_sheet(wb):
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    body_font = Font(size=10)
    muted_font = Font(size=9, color="666666")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="CCCCCC")
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F0F4F8")

    ws["A1"] = "くらしの計算室 — 手取り試算スプレッドシート（2026年版）"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")

    ws["A3"] = "免責事項"
    ws["A3"].font = section_font
    disclaimers = [
        "本テンプレートの計算結果はすべて概算です。最終的な手取り額は給与明細・勤務先の人事・税理士等でご確認ください。",
        "健康保険は協会けんぽの標準料率（本人4.95%）を目安にしています。健保組合・扶養家族・標準報酬月額の等級差は反映しません。",
        "住民税の天引き開始時期（新卒1年目の非課税期間など）や、各種控除（配偶者・住宅ローン等）は未対応です。",
        "ボーナスにも社会保険料率（合計14.65%）を適用します。サイトの無料ツール tools/tedori.html と同一ロジックです。",
        "税制・社保料率は改正により変わります。毎年8月頃は料率の見直しを推奨します。",
    ]
    for i, text in enumerate(disclaimers, start=4):
        ws[f"A{i}"] = f"・{text}"
        ws[f"A{i}"].font = body_font
        ws[f"A{i}"].alignment = wrap
        ws.merge_cells(f"A{i}:D{i}")

    row = 10
    ws[f"A{row}"] = "使い方"
    ws[f"A{row}"].font = section_font
    usage = [
        ("1. 手取り試算シート", "B3（月収）・B4（年間ボーナス）を入力すると、社保・税の内訳と毎月の手取り（B17）が自動計算されます。"),
        ("2. 手取り比較シート", "パターンA〜Cの月収・ボーナスを入力し、手取りとAとの差額（月・年）を横並びで比較できます。転職・昇給の検討に使えます。"),
        ("3. 家計サマリシート", "手取り試算の手取りを参照し、固定費・変動費を入力すると「自由に使える額」と貯蓄率がわかります。各費目は自由に書き換えてください。"),
        ("4. 積立メモシート", "家計サマリの余剰額を積立額の初期値に、想定利回り・期間から将来の資産額を複利で試算します（tsumitate.html と同一ロジック）。"),
        ("5. 保存・編集", "数値は自由に書き換えてください。別名で保存すれば、複数シナリオをファイルごとに残せます。"),
        ("6. 推奨環境", "編集はPC版 Excel または Google スプレッドシート（xlsxインポート）を推奨します。スマホは閲覧のみ想定です。"),
    ]
    for title, detail in usage:
        row += 1
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, size=10)
        ws[f"B{row}"] = detail
        ws[f"B{row}"].font = body_font
        ws[f"B{row}"].alignment = wrap
        ws.merge_cells(f"B{row}:D{row}")

    row += 2
    ws[f"A{row}"] = "検証用テストケース"
    ws[f"A{row}"].font = section_font
    row += 1
    headers = ("月収（額面）", "年間ボーナス", "毎月手取り（目安）", "確認方法")
    for col, header in zip("ABCD", headers):
        cell = ws[f"{col}{row}"]
        cell.value = header
        cell.font = Font(bold=True, size=10)
        cell.fill = header_fill
        cell.border = table_border
        cell.alignment = Alignment(horizontal="center")
    cases = [
        (300_000, 0, 237_184, "手取り比較シートのパターンA初期値"),
        (330_000, 0, 259_827, "手取り比較シートのパターンB初期値"),
        (360_000, 0, 282_471, "手取り比較シートのパターンC初期値"),
    ]
    for salary, bonus, net, note in cases:
        row += 1
        ws[f"A{row}"] = salary
        ws[f"B{row}"] = bonus
        ws[f"C{row}"] = net
        ws[f"D{row}"] = note
        ws[f"A{row}"].number_format = "#,##0"
        ws[f"B{row}"].number_format = "#,##0"
        ws[f"C{row}"].number_format = "#,##0"
        for col in "ABCD":
            ws[f"{col}{row}"].font = body_font
            ws[f"{col}{row}"].border = table_border
        ws[f"D{row}"].alignment = wrap

    row += 2
    ws[f"A{row}"] = "無料Webツール"
    ws[f"A{row}"].font = section_font
    row += 1
    ws[f"A{row}"] = "https://shunsukesaito00.github.io/kurashi/tools/tedori.html"
    ws[f"A{row}"].font = Font(size=10, color="0563C1", underline="single")
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    ws[f"A{row}"] = "ブラウザで同じ条件を試算できます。共有URLで条件を保存・共有することも可能です。"
    ws[f"A{row}"].font = muted_font
    ws.merge_cells(f"A{row}:D{row}")


def build_single_sheet(wb):
    ws = wb.create_sheet("手取り試算")
    style_header(ws, "手取り試算 — くらしの計算室（2026年版）")

    labels = {
        3: "月収（額面・円）",
        4: "年間ボーナス（円）",
        5: "年間額面",
        6: "健康保険料（月）",
        7: "厚生年金（月）",
        8: "雇用保険（月）",
        9: "社会保険料（年）",
        10: "給与所得控除（年）",
        11: "課税所得・所得税用（年）",
        12: "所得税（年・復興含む）",
        13: "課税標準・住民税用（年）",
        14: "住民税（年）",
        15: "所得税（月割）",
        16: "住民税（月割）",
        17: "毎月の手取り",
        18: "控除合計（月）",
        19: "手取り率",
        20: "年間手取り目安（簡易）",
    }
    for row, label in labels.items():
        ws[f"A{row}"] = label
        if row >= 5:
            ws[f"A{row}"].font = Font(color="444444")

    ws["B3"] = 300000
    ws["B4"] = 0
    apply_net_formulas(ws, "B")
    ws["B18"] = "=B3-B17"
    ws["B19"] = "=IF(B3=0,\"\",B17/B3)"
    ws["B19"].number_format = "0.0%"
    ws["B20"] = "=B17*12"

    style_yen(
        ws,
        [f"B{r}" for r in range(3, 21) if r != 19],
    )
    ws["B17"].font = Font(bold=True, size=12)
    ws["B17"].fill = PatternFill("solid", fgColor="E8F4EA")


def build_compare_sheet(wb):
    ws = wb.create_sheet("手取り比較")
    style_header(ws, "手取り比較（最大3パターン）— くらしの計算室")

    ws["B2"] = "パターンA"
    ws["C2"] = "パターンB"
    ws["D2"] = "パターンC"
    for c in "BCD":
        ws[f"{c}2"].font = Font(bold=True)
        ws[f"{c}2"].alignment = Alignment(horizontal="center")

    ws["A3"] = "月収（額面・円）"
    ws["A4"] = "年間ボーナス（円）"
    mid_labels = {
        5: "年間額面",
        6: "健康保険料（月）",
        7: "厚生年金（月）",
        8: "雇用保険（月）",
        9: "社会保険料（年）",
        10: "給与所得控除（年）",
        11: "課税所得・所得税用（年）",
        12: "所得税（年・復興含む）",
        13: "課税標準・住民税用（年）",
        14: "住民税（年）",
        15: "所得税（月割）",
        16: "住民税（月割）",
    }
    for row, label in mid_labels.items():
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(color="444444")
    ws["B3"], ws["C3"], ws["D3"] = 300000, 330000, 360000
    ws["B4"], ws["C4"], ws["D4"] = 0, 0, 0

    for col in "BCD":
        apply_net_formulas(ws, col)

    summary = {
        17: "毎月の手取り",
        18: "手取り率",
        19: "Aとの手取り差（月）",
        20: "Aとの手取り差（年）",
    }
    for row, label in summary.items():
        ws[f"A{row}"] = label

    for col in "BCD":
        ws[f"{col}18"] = f"=IF({col}3=0,\"\",{col}17/{col}3)"
        ws[f"{col}18"].number_format = "0.0%"
    ws["B19"] = "—"
    ws["B20"] = "—"
    ws["C19"] = "=$C$17-$B$17"
    ws["D19"] = "=$D$17-$B$17"
    ws["C20"] = "=C19*12"
    ws["D20"] = "=D19*12"

    for col in "BCD":
        ws[f"{col}17"].font = Font(bold=True)
        style_yen(ws, [f"{col}{r}" for r in range(3, 21) if r not in (18, 19) or r == 17])
    ws["C19"].number_format = "#,##0"
    ws["D19"].number_format = "#,##0"
    ws["C20"].number_format = "#,##0"
    ws["D20"].number_format = "#,##0"


def build_budget_summary_sheet(wb):
    ws = wb.create_sheet("家計サマリ")
    style_header(ws, "家計サマリ — くらしの計算室")

    section_font = Font(bold=True, size=11, color="333333")
    label_font = Font(size=10)
    result_fill = PatternFill("solid", fgColor="E8F4EA")

    ws["A3"] = "毎月の手取り（手取り試算より）"
    ws["B3"] = "='手取り試算'!B17"
    ws["A3"].font = label_font
    ws["B3"].font = Font(bold=True, size=11)
    ws["B3"].number_format = "#,##0"

    ws["A5"] = "固定費（毎月）"
    ws["A5"].font = section_font
    fixed_items = [
        (6, "家賃・住宅ローン", 80_000),
        (7, "水道光熱費", 12_000),
        (8, "通信費（携帯・ネット）", 8_000),
        (9, "保険（生命・医療等）", 5_000),
        (10, "ローン返済（住宅以外）", 0),
        (11, "その他固定費", 5_000),
    ]
    for row, label, value in fixed_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = label_font
        ws[f"B{row}"] = value
        ws[f"B{row}"].number_format = "#,##0"

    ws["A12"] = "固定費合計"
    ws["A12"].font = Font(bold=True)
    ws["B12"] = "=SUM(B6:B11)"
    ws["B12"].font = Font(bold=True)
    ws["B12"].number_format = "#,##0"

    ws["A14"] = "変動費（毎月）"
    ws["A14"].font = section_font
    variable_items = [
        (15, "食費", 40_000),
        (16, "交通費", 8_000),
        (17, "日用品・消耗品", 5_000),
        (18, "交際費・娯楽", 15_000),
        (19, "その他変動費", 5_000),
    ]
    for row, label, value in variable_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = label_font
        ws[f"B{row}"] = value
        ws[f"B{row}"].number_format = "#,##0"

    ws["A20"] = "変動費合計"
    ws["A20"].font = Font(bold=True)
    ws["B20"] = "=SUM(B15:B19)"
    ws["B20"].font = Font(bold=True)
    ws["B20"].number_format = "#,##0"

    ws["A22"] = "支出合計（固定費＋変動費）"
    ws["A22"].font = label_font
    ws["B22"] = "=B12+B20"
    ws["B22"].number_format = "#,##0"

    ws["A23"] = "自由に使える額（余剰）"
    ws["A23"].font = Font(bold=True, size=11)
    ws["B23"] = "=B3-B22"
    ws["B23"].font = Font(bold=True, size=12)
    ws["B23"].fill = result_fill
    ws["B23"].number_format = "#,##0"

    ws["A24"] = "貯蓄に回せる割合（手取り比）"
    ws["A24"].font = label_font
    ws["B24"] = '=IF(B3=0,"",B23/B3)'
    ws["B24"].number_format = "0.0%"

    ws["A25"] = "支出率（手取り比）"
    ws["A25"].font = label_font
    ws["B25"] = '=IF(B3=0,"",B22/B3)'
    ws["B25"].number_format = "0.0%"

    ws["A27"] = "メモ"
    ws["A27"].font = section_font
    ws["A28"] = "自由に使える額は、手取りから固定費・変動費を引いた残りです。実際の貯蓄額は積立・投資の振替額で調整してください。"
    ws["A28"].font = Font(size=9, color="666666")
    ws.merge_cells("A28:B28")
    ws["A28"].alignment = Alignment(wrap_text=True, vertical="top")


def tsumitate_total_formula(monthly_cell: str, rate_cell: str, years_cell: str) -> str:
    """月次複利の積立終価（tsumitate.html の calc と同一）。"""
    r = f"{rate_cell}/100/12"
    n = f"({years_cell}*12)"
    return (
        f"=IF({years_cell}<=0,0,"
        f"IF({rate_cell}=0,ROUND({monthly_cell}*{n},0),"
        f"ROUND({monthly_cell}*(((1+{r})^{n}-1)/{r}),0)))"
    )


def build_savings_sheet(wb):
    ws = wb.create_sheet("積立メモ")
    style_header(ws, "積立メモ — くらしの計算室")

    section_font = Font(bold=True, size=11, color="333333")
    label_font = Font(size=10)
    result_fill = PatternFill("solid", fgColor="E8F4EA")
    thin = Side(style="thin", color="CCCCCC")
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F0F4F8")

    ws["A3"] = "毎月の積立額（円）"
    ws["B3"] = "=MAX(0,'家計サマリ'!B23)"
    ws["A4"] = "想定利回り（年率%）"
    ws["B4"] = 5
    ws["A5"] = "積立期間（年）"
    ws["B5"] = 20
    for row in (3, 4, 5):
        ws[f"A{row}"].font = label_font
    ws["B3"].number_format = "#,##0"
    ws["B4"].number_format = "0.0"

    ws["A7"] = "シミュレーション結果"
    ws["A7"].font = section_font
    ws["A8"] = "最終的な資産額（目安）"
    ws["B8"] = tsumitate_total_formula("B3", "B4", "B5")
    ws["A9"] = "積立元本"
    ws["B9"] = "=B3*B5*12"
    ws["A10"] = "運用益"
    ws["B10"] = "=B8-B9"
    ws["A11"] = "元本に対する増加率"
    ws["B11"] = '=IF(B9=0,"",B10/B9)'
    ws["B11"].number_format = "0.0%"
    for row in (8, 9, 10):
        ws[f"A{row}"].font = label_font
        ws[f"B{row}"].number_format = "#,##0"
    ws["B8"].font = Font(bold=True, size=12)
    ws["B8"].fill = result_fill

    ws["A13"] = "期間別の目安（上記の積立額・利回りで試算）"
    ws["A13"].font = section_font
    row = 14
    for col, header in zip("ABC", ("期間（年）", "積立元本", "資産額（目安）")):
        cell = ws[f"{col}{row}"]
        cell.value = header
        cell.font = Font(bold=True, size=10)
        cell.fill = header_fill
        cell.border = table_border
        cell.alignment = Alignment(horizontal="center")
    for years in (10, 20, 30):
        row += 1
        ws[f"A{row}"] = years
        ws[f"B{row}"] = f"=$B$3*A{row}*12"
        ws[f"C{row}"] = tsumitate_total_formula("$B$3", "$B$4", f"A{row}")
        for col in "ABC":
            ws[f"{col}{row}"].font = label_font
            ws[f"{col}{row}"].border = table_border
        ws[f"B{row}"].number_format = "#,##0"
        ws[f"C{row}"].number_format = "#,##0"

    ws["A19"] = "無料Webツール"
    ws["A19"].font = section_font
    ws["A20"] = "https://shunsukesaito00.github.io/kurashi/tools/tsumitate.html"
    ws["A20"].font = Font(size=10, color="0563C1", underline="single")
    ws.merge_cells("A20:B20")

    ws["A22"] = "免責"
    ws["A22"].font = section_font
    ws["A23"] = "月次複利・利回り一定を仮定した概算です。実際の運用成果を保証するものではなく、金融商品には価格変動リスクがあります。"
    ws["A23"].font = Font(size=9, color="666666")
    ws.merge_cells("A23:B23")
    ws["A23"].alignment = Alignment(wrap_text=True, vertical="top")


def verify_with_tedori_logic():
    """tools/tedori.html の computeNet と代表例が一致するか（Python再実装）。"""

    def salary_deduction(gross):
        if gross <= 1_625_000:
            return 550_000
        if gross <= 1_800_000:
            return gross * 0.4 - 100_000
        if gross <= 3_600_000:
            return gross * 0.3 + 80_000
        if gross <= 6_600_000:
            return gross * 0.2 + 440_000
        if gross <= 8_500_000:
            return gross * 0.1 + 1_100_000
        return 1_950_000

    def income_tax_annual(taxable):
        brackets = [
            (1_950_000, 0.05, 0),
            (3_300_000, 0.10, 97_500),
            (6_950_000, 0.20, 427_500),
            (9_000_000, 0.23, 636_000),
            (18_000_000, 0.33, 1_536_000),
            (40_000_000, 0.40, 2_796_000),
            (float("inf"), 0.45, 4_796_000),
        ]
        if taxable <= 0:
            return 0
        for limit, rate, ded in brackets:
            if taxable <= limit:
                return taxable * rate - ded
        return 0

    def compute_net(salary, bonus=0):
        annual = salary * 12 + bonus
        health = salary * 0.0495
        pension = salary * 0.0915
        emp = salary * 0.0055
        social = (health + pension + emp) * 12 + bonus * 0.1465
        taxable = max(0, annual - salary_deduction(annual) - social - 480_000)
        income = income_tax_annual(taxable) * 1.021
        res_taxable = max(0, annual - salary_deduction(annual) - social - 430_000)
        resident = res_taxable * 0.10 + 5_000
        monthly_income = income / 12 * (salary * 12 / annual) if annual else 0
        return round(salary - health - pension - emp - monthly_income - resident / 12)

    cases = [(300_000, 237_184), (330_000, 259_827), (360_000, 282_471)]
    for salary, expected in cases:
        got = compute_net(salary)
        if got != expected:
            raise SystemExit(f"検証失敗: 月収{salary} → {got} (期待 {expected})")


def verify_tsumitate_logic():
    """tools/tsumitate.html の calc と代表例が一致するか。"""

    def calc_total(monthly, rate, years):
        if years <= 0:
            return 0
        r = rate / 100 / 12
        n = years * 12
        if r == 0:
            return round(monthly * n)
        return round(monthly * ((1 + r) ** n - 1) / r)

    cases = [
        (30_000, 5, 10, 4_658_468),
        (30_000, 5, 20, 12_331_010),
        (30_000, 5, 30, 24_967_759),
    ]
    for monthly, rate, years, expected in cases:
        got = calc_total(monthly, rate, years)
        if got != expected:
            raise SystemExit(
                f"積立検証失敗: {monthly}円・{rate}%・{years}年 → {got} (期待 {expected})"
            )


def main():
    verify_with_tedori_logic()
    verify_tsumitate_logic()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_readme_sheet(wb)
    build_single_sheet(wb)
    build_compare_sheet(wb)
    build_budget_summary_sheet(wb)
    build_savings_sheet(wb)
    wb.save(OUT)
    print(f"OK: {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
